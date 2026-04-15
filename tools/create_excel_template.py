#!/usr/bin/env python3
"""Generate the blank Excel template for CMDB-driven asset transfers.

Run once to create the template:
    python tools/create_excel_template.py

Produces:
    tools/asset_transfer_template.xlsx
    tools/asset_transfer_sample.xlsx   (pre-filled with example rows)
"""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    raise SystemExit(
        "openpyxl is required.  Install it:\n"
        "  pip install openpyxl\n"
    )

OUT_DIR = Path(__file__).resolve().parent

# ---------- column definitions ----------
COLUMNS = [
    # col_name,                  width, description
    ("REQUEST_ID",               18,    "Unique ID for this transfer request (e.g. CMDB_EVT_0001)"),
    ("ASSET_NUMBER",             18,    "Oracle FA asset number (e.g. 142847)"),
    ("ASSET_ID",                 14,    "Oracle FA asset ID (numeric). Leave blank to auto-lookup."),
    ("BOOK_TYPE_CODE",           22,    "Source FA book (e.g. US CORP BOOK)"),
    ("TRANSFER_TO_ENTITY",       22,    "Target entity name (e.g. UK Entity)"),
    ("TRANSFER_DATE",            16,    "Transfer effective date (YYYY-MM-DD)"),
    ("TRANSFER_TO_LOCATION",     24,    "Target location text / DFF value"),
    ("TARGET_LOCATION_ID",       20,    "Target location CCID (numeric)"),
    ("TARGET_EXPENSE_CCID",      22,    "Target expense CCID (numeric, optional)"),
    ("TARGET_BOOK_TYPE_CODE",    24,    "Target FA book (e.g. UK CORP BOOK). Leave blank to resolve from entity."),
    ("RAW_TARGET_COST_CENTER",   24,    "Raw cost center DFF value (optional)"),
    ("NOTES",                    30,    "Free-text notes (not sent to Oracle)"),
]

SAMPLE_ROWS = [
    {
        "REQUEST_ID": "CMDB_EVT_0001",
        "ASSET_NUMBER": "142847",
        "ASSET_ID": "",
        "BOOK_TYPE_CODE": "US CORP BOOK",
        "TRANSFER_TO_ENTITY": "UK Entity",
        "TRANSFER_DATE": "2026-03-31",
        "TRANSFER_TO_LOCATION": "LON-DC-01",
        "TARGET_LOCATION_ID": "300100888001",
        "TARGET_EXPENSE_CCID": "627564",
        "TARGET_BOOK_TYPE_CODE": "UK CORP BOOK",
        "RAW_TARGET_COST_CENTER": "20501",
        "NOTES": "Server relocation London",
    },
    {
        "REQUEST_ID": "CMDB_EVT_0002",
        "ASSET_NUMBER": "142848",
        "ASSET_ID": "",
        "BOOK_TYPE_CODE": "US CORP BOOK",
        "TRANSFER_TO_ENTITY": "JP Entity",
        "TRANSFER_DATE": "2026-03-31",
        "TRANSFER_TO_LOCATION": "TOK-DC-02",
        "TARGET_LOCATION_ID": "300100888010",
        "TARGET_EXPENSE_CCID": "",
        "TARGET_BOOK_TYPE_CODE": "JP CORP BOOK",
        "RAW_TARGET_COST_CENTER": "30501",
        "NOTES": "Switch relocation Tokyo",
    },
    {
        "REQUEST_ID": "CMDB_EVT_0003",
        "ASSET_NUMBER": "142850",
        "ASSET_ID": "300100505",
        "BOOK_TYPE_CODE": "US CORP BOOK",
        "TRANSFER_TO_ENTITY": "US Entity",
        "TRANSFER_DATE": "2026-04-01",
        "TRANSFER_TO_LOCATION": "",
        "TARGET_LOCATION_ID": "300100999001",
        "TARGET_EXPENSE_CCID": "",
        "TARGET_BOOK_TYPE_CODE": "US CORP BOOK",
        "RAW_TARGET_COST_CENTER": "",
        "NOTES": "Same-book location change",
    },
    {
        "REQUEST_ID": "CMDB_EVT_0004",
        "ASSET_NUMBER": "142851",
        "ASSET_ID": "",
        "BOOK_TYPE_CODE": "US CORP BOOK",
        "TRANSFER_TO_ENTITY": "DE Entity",
        "TRANSFER_DATE": "2026-04-01",
        "TRANSFER_TO_LOCATION": "FRA-DC-01",
        "TARGET_LOCATION_ID": "300100888007",
        "TARGET_EXPENSE_CCID": "630200",
        "TARGET_BOOK_TYPE_CODE": "DE CORP BOOK",
        "RAW_TARGET_COST_CENTER": "40501",
        "NOTES": "Storage array to Frankfurt",
    },
]


def _build_workbook(include_samples: bool = False) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Asset Transfers"

    # --- Styles ---
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # --- Header row ---
    required_cols = {"REQUEST_ID", "ASSET_NUMBER", "BOOK_TYPE_CODE", "TRANSFER_TO_ENTITY", "TRANSFER_DATE"}
    for col_idx, (col_name, width, _desc) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Description row (row 2) ---
    desc_font = Font(name="Calibri", italic=True, size=9, color="666666")
    for col_idx, (_col_name, _width, desc) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = desc_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border

    # --- Highlight required columns (light yellow) for data rows ---
    if include_samples:
        for row_idx, sample in enumerate(SAMPLE_ROWS, start=3):
            for col_idx, (col_name, _w, _d) in enumerate(COLUMNS, start=1):
                val = sample.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if col_name in required_cols:
                    cell.fill = required_fill
    else:
        # Add 20 empty rows with formatting
        for row_idx in range(3, 23):
            for col_idx, (col_name, _w, _d) in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value="")
                cell.border = thin_border
                if col_name in required_cols:
                    cell.fill = required_fill

    # --- Data validation for TRANSFER_DATE (YYYY-MM-DD hint) ---
    # Freeze header
    ws.freeze_panes = "A3"

    # --- Instructions sheet ---
    instr = wb.create_sheet("Instructions")
    instructions = [
        ("OFAM Asset Transfer Excel Template", None),
        ("", None),
        ("Required columns (yellow highlight):", None),
        ("  REQUEST_ID", "Unique identifier for each transfer request"),
        ("  ASSET_NUMBER", "Oracle Fixed Assets asset number"),
        ("  BOOK_TYPE_CODE", "Source FA book (e.g. 'US CORP BOOK')"),
        ("  TRANSFER_TO_ENTITY", "Target entity (must match entity_book_map in config)"),
        ("  TRANSFER_DATE", "Effective date in YYYY-MM-DD format"),
        ("", None),
        ("Optional columns:", None),
        ("  ASSET_ID", "If blank, looked up via getAssetInformation API"),
        ("  TRANSFER_TO_LOCATION", "Location text / DFF value"),
        ("  TARGET_LOCATION_ID", "Numeric location CCID"),
        ("  TARGET_EXPENSE_CCID", "Numeric expense CCID"),
        ("  TARGET_BOOK_TYPE_CODE", "If blank, resolved from TRANSFER_TO_ENTITY"),
        ("  RAW_TARGET_COST_CENTER", "Cost center DFF value"),
        ("  NOTES", "For your reference only, not sent to Oracle"),
        ("", None),
        ("Usage:", None),
        ("  python tools/run_excel_transfers.py --excel tools/asset_transfer_sample.xlsx --dry-run", None),
        ("  python tools/run_excel_transfers.py --excel tools/asset_transfer_sample.xlsx --execute", None),
        ("", None),
        ("Row 2 contains column descriptions. Data starts at row 3.", None),
    ]
    for row_idx, (col_a, col_b) in enumerate(instructions, start=1):
        cell = instr.cell(row=row_idx, column=1, value=col_a)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        if col_b:
            instr.cell(row=row_idx, column=2, value=col_b)
    instr.column_dimensions["A"].width = 40
    instr.column_dimensions["B"].width = 60

    return wb


def main() -> None:
    # Blank template
    wb_blank = _build_workbook(include_samples=False)
    blank_path = OUT_DIR / "asset_transfer_template.xlsx"
    wb_blank.save(str(blank_path))
    print(f"Created blank template: {blank_path}")

    # Sample with data
    wb_sample = _build_workbook(include_samples=True)
    sample_path = OUT_DIR / "asset_transfer_sample.xlsx"
    wb_sample.save(str(sample_path))
    print(f"Created sample file:    {sample_path}")


if __name__ == "__main__":
    main()
