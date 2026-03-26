#!/usr/bin/env python3
"""Read an Excel spreadsheet of CMDB asset transfer requests and execute them
against Oracle Fusion Fixed Assets via the ERP Integrations REST API.

This script bridges the gap between a CMDB-style Excel input and the existing
OFAM transfer engine.  It supports two modes:

  1. **DFF-only mode** (default, ``--mode dff``)
     Calls ``processTransaction-updateAssetDescriptiveDetails`` for each row.
     This stamps the DFF values on the asset so the downstream BIP report
     picks them up and the existing transfer engine processes them.

  2. **Full transfer mode** (``--mode transfer``)
     Calls ``getAssetInformation`` to read current state, then builds and
     submits a ``transferAsset`` or ``bookTransfer`` payload directly.

Examples
--------
  # Dry-run (no Oracle calls, just validate + build payloads)
  python tools/run_excel_transfers.py \\
      --excel tools/asset_transfer_sample.xlsx \\
      --config sample_bip_with_pre_dff_updates.json \\
      --dry-run

  # Execute DFF updates against dev
  python tools/run_excel_transfers.py \\
      --excel tools/asset_transfer_sample.xlsx \\
      --config sample_bip_with_pre_dff_updates.json \\
      --execute

  # Full transfer mode (get asset info + transfer)
  python tools/run_excel_transfers.py \\
      --excel tools/asset_transfer_sample.xlsx \\
      --config sample_bip_with_pre_dff_updates.json \\
      --mode transfer \\
      --execute
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit(
        "openpyxl is required.  Install it:\n"
        "  pip install openpyxl\n"
    )

REPO_ROOT = Path(__file__).resolve().parents[1]
SRC_ROOT = REPO_ROOT / "src" / "main" / "python"
if str(SRC_ROOT) not in sys.path:
    sys.path.insert(0, str(SRC_ROOT))

from ofam_asset_xfer.exceptions import ConfigError, FusionApiError, ValidationError
from ofam_asset_xfer.oracle_client import OracleConfig, OracleErpIntegrationsClient
from ofam_asset_xfer.paramlist import build_parameter_list
from ofam_asset_xfer.pre_bip_dff import (
    DffUpdateRequest,
    DffUpdateResult,
    PreBipDffConfig,
    PreBipDffUpdater,
)
from ofam_asset_xfer.fusion_ops import (
    AssetState,
    get_asset_information,
    build_same_book_transfer_params,
    build_book_transfer_params,
)
from ofam_asset_xfer.entity_resolver import EntityBookResolver

log = logging.getLogger(__name__)

# ---------- Excel reading ----------

REQUIRED_COLUMNS = {"REQUEST_ID", "ASSET_NUMBER", "BOOK_TYPE_CODE", "TRANSFER_TO_ENTITY", "TRANSFER_DATE"}

DESCRIPTION_ROW = 2  # Row 2 has column descriptions, data starts at row 3
DATA_START_ROW = 3


@dataclass
class ExcelRow:
    """One parsed row from the Excel template."""
    row_number: int
    request_id: str
    asset_number: str
    asset_id: str
    book_type_code: str
    transfer_to_entity: str
    transfer_date: str
    transfer_to_location: str
    target_location_id: str
    target_expense_ccid: str
    target_book_type_code: str
    raw_target_cost_center: str
    account: str
    notes: str

    def to_dff_update_request(self) -> DffUpdateRequest:
        """Convert to a DffUpdateRequest for the pre-BIP DFF flow."""
        fields: Dict[str, Any] = {}
        if self.transfer_to_entity:
            fields["transfer_to_entity"] = self.transfer_to_entity
        if self.transfer_date:
            fields["transfer_date"] = self.transfer_date
        if self.transfer_to_location:
            fields["transfer_to_location"] = self.transfer_to_location
        if self.raw_target_cost_center:
            fields["raw_target_cost_center"] = self.raw_target_cost_center
        if self.account:
            fields["account"] = self.account

        return DffUpdateRequest(
            request_id=self.request_id,
            asset_number=self.asset_number,
            book_type_code=self.book_type_code,
            fields=fields,
            asset_id=self.asset_id or None,
        )


def read_excel(path: Path) -> List[ExcelRow]:
    """Read the Asset Transfers sheet and return parsed rows."""
    wb = load_workbook(str(path), read_only=True, data_only=True)

    # Find the right sheet
    sheet_name = "Asset Transfers"
    if sheet_name not in wb.sheetnames:
        # Fallback: use the first sheet
        sheet_name = wb.sheetnames[0]
        log.warning("Sheet 'Asset Transfers' not found, using '%s'", sheet_name)

    ws = wb[sheet_name]

    # Read header row (row 1)
    headers: List[str] = []
    for cell in ws[1]:
        val = str(cell.value or "").strip().upper()
        headers.append(val)

    # Validate required columns
    header_set = set(headers)
    missing = REQUIRED_COLUMNS - header_set
    if missing:
        raise ConfigError(
            f"Excel is missing required columns: {sorted(missing)}. "
            f"Found columns: {sorted(header_set)}"
        )

    col_index = {name: idx for idx, name in enumerate(headers)}

    def _get(row_cells: tuple, col_name: str) -> str:
        idx = col_index.get(col_name)
        if idx is None or idx >= len(row_cells):
            return ""
        val = row_cells[idx].value
        if val is None:
            return ""
        # Handle date objects from Excel
        if isinstance(val, (date, datetime)):
            return val.strftime("%Y-%m-%d")
        return str(val).strip()

    rows: List[ExcelRow] = []
    for row_idx, row_cells in enumerate(ws.iter_rows(min_row=DATA_START_ROW), start=DATA_START_ROW):
        request_id = _get(row_cells, "REQUEST_ID")
        if not request_id:
            continue  # Skip empty rows

        asset_number = _get(row_cells, "ASSET_NUMBER")
        if not asset_number:
            log.warning("Row %d: REQUEST_ID=%s has no ASSET_NUMBER, skipping", row_idx, request_id)
            continue

        rows.append(
            ExcelRow(
                row_number=row_idx,
                request_id=request_id,
                asset_number=asset_number,
                asset_id=_get(row_cells, "ASSET_ID"),
                book_type_code=_get(row_cells, "BOOK_TYPE_CODE"),
                transfer_to_entity=_get(row_cells, "TRANSFER_TO_ENTITY"),
                transfer_date=_get(row_cells, "TRANSFER_DATE"),
                transfer_to_location=_get(row_cells, "TRANSFER_TO_LOCATION"),
                target_location_id=_get(row_cells, "TARGET_LOCATION_ID"),
                target_expense_ccid=_get(row_cells, "TARGET_EXPENSE_CCID"),
                target_book_type_code=_get(row_cells, "TARGET_BOOK_TYPE_CODE"),
                raw_target_cost_center=_get(row_cells, "RAW_TARGET_COST_CENTER"),
                account=_get(row_cells, "ACCOUNT"),
                notes=_get(row_cells, "NOTES"),
            )
        )

    wb.close()
    return rows


# ---------- Transfer result ----------

@dataclass
class TransferResult:
    request_id: str
    asset_number: str
    book_type_code: str
    mode: str  # dff, transfer
    status: str  # UPDATED, DRY_RUN, TRANSFERRED, FAILED, NOOP
    asset_id: str = ""
    target_book: str = ""
    error: str = ""
    fusion_response: Optional[Dict[str, Any]] = None
    planned_params: Optional[Dict[str, Any]] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "request_id": self.request_id,
            "asset_number": self.asset_number,
            "book_type_code": self.book_type_code,
            "mode": self.mode,
            "status": self.status,
            "asset_id": self.asset_id,
            "target_book": self.target_book,
            "error": self.error,
            "fusion_response": self.fusion_response,
            "planned_params": self.planned_params,
        }


# ---------- DFF mode ----------

def run_dff_mode(
    rows: List[ExcelRow],
    config: Dict[str, Any],
    dry_run: bool,
) -> List[TransferResult]:
    """Update DFF values on each asset via processTransaction-updateAssetDescriptiveDetails."""
    results: List[TransferResult] = []

    # Build PreBipDffConfig from config file
    pre_bip_block = config.get("pre_bip_dff_updates", {})
    dff_cfg = PreBipDffConfig.from_dict({**pre_bip_block, "enabled": True})

    if dry_run:
        # Dry-run: just build payloads without calling Oracle
        for row in rows:
            req = row.to_dff_update_request()
            # Build params manually to show what would be sent
            params: Dict[str, Any] = dict(dff_cfg.static_params)
            params["P_ASSET_ID"] = row.asset_id or f"<lookup:{row.asset_number}>"
            if dff_cfg.request_id_param:
                params[dff_cfg.request_id_param] = row.request_id
            if dff_cfg.trace_param:
                params[dff_cfg.trace_param] = dff_cfg.trace_value

            # Map logical fields to Fusion params
            for logical_name, fusion_param in dff_cfg.field_parameter_map.items():
                val = req.fields.get(logical_name)
                if val is not None and str(val).strip():
                    params[fusion_param] = val

            # Also include book_type_code if mapped
            if "book_type_code" in dff_cfg.field_parameter_map:
                params[dff_cfg.field_parameter_map["book_type_code"]] = row.book_type_code

            results.append(
                TransferResult(
                    request_id=row.request_id,
                    asset_number=row.asset_number,
                    book_type_code=row.book_type_code,
                    mode="dff",
                    status="DRY_RUN",
                    asset_id=row.asset_id,
                    planned_params=params,
                )
            )
        return results

    # Execute: create Oracle client and run updates
    oracle_cfg = OracleConfig.from_dict(config["oracle"])
    client = OracleErpIntegrationsClient(oracle_cfg)
    updater = PreBipDffUpdater(client, dff_cfg)

    for row in rows:
        req = row.to_dff_update_request()
        try:
            result = updater.apply_update(req, dry_run=False)
            results.append(
                TransferResult(
                    request_id=result.request_id,
                    asset_number=result.asset_number,
                    book_type_code=result.book_type_code,
                    mode="dff",
                    status=result.status,
                    asset_id=result.asset_id or "",
                    error=result.error or "",
                    fusion_response=result.fusion_response,
                )
            )
        except Exception as exc:
            results.append(
                TransferResult(
                    request_id=row.request_id,
                    asset_number=row.asset_number,
                    book_type_code=row.book_type_code,
                    mode="dff",
                    status="FAILED",
                    error=str(exc),
                )
            )

    return results


# ---------- Full transfer mode ----------

def run_transfer_mode(
    rows: List[ExcelRow],
    config: Dict[str, Any],
    dry_run: bool,
) -> List[TransferResult]:
    """Full transfer: getAssetInformation + transferAsset/bookTransfer."""
    results: List[TransferResult] = []

    entity_map = config.get("entity_book_map", {})
    resolver = EntityBookResolver(entity_map)

    if dry_run:
        # Build planned payloads without calling Oracle
        for row in rows:
            target_book = row.target_book_type_code or resolver.resolve_target_book(row.transfer_to_entity)
            is_cross_book = target_book.strip().upper() != row.book_type_code.strip().upper()

            results.append(
                TransferResult(
                    request_id=row.request_id,
                    asset_number=row.asset_number,
                    book_type_code=row.book_type_code,
                    mode="transfer",
                    status="DRY_RUN",
                    asset_id=row.asset_id,
                    target_book=target_book,
                    planned_params={
                        "transfer_type": "cross-book (bookTransfer)" if is_cross_book else "same-book (transferAsset)",
                        "target_book": target_book,
                        "target_location_id": row.target_location_id,
                        "target_expense_ccid": row.target_expense_ccid,
                        "transfer_date": row.transfer_date,
                        "note": "Actual params built after getAssetInformation call",
                    },
                )
            )
        return results

    # Execute: create Oracle client
    oracle_cfg = OracleConfig.from_dict(config["oracle"])
    client = OracleErpIntegrationsClient(oracle_cfg)

    for row in rows:
        try:
            # Step 1: Resolve target book
            target_book = row.target_book_type_code or resolver.resolve_target_book(row.transfer_to_entity)
            is_cross_book = target_book.strip().upper() != row.book_type_code.strip().upper()

            # Step 2: Get current asset state
            _raw, _pl, state = get_asset_information(
                client,
                row.book_type_code,
                row.asset_number,
                asset_id=row.asset_id or None,
            )

            # Step 3: Build overrides from Excel row
            overrides: Dict[str, Any] = {}
            if row.target_location_id:
                overrides["location_ccid"] = row.target_location_id
            if row.target_expense_ccid:
                overrides["expense_ccid"] = row.target_expense_ccid

            # Step 4: Build and submit transfer
            if is_cross_book:
                params = build_book_transfer_params(
                    state=state,
                    dest_book_type_code=target_book,
                    effective_date=row.transfer_date,
                    overrides=overrides,
                    request_id=row.request_id,
                )
                raw, pl = client.process_transaction("transferAsset", params)
            else:
                params, is_noop = build_same_book_transfer_params(
                    state=state,
                    effective_date=row.transfer_date,
                    overrides=overrides,
                    request_id=row.request_id,
                )
                if is_noop:
                    results.append(
                        TransferResult(
                            request_id=row.request_id,
                            asset_number=row.asset_number,
                            book_type_code=row.book_type_code,
                            mode="transfer",
                            status="NOOP",
                            asset_id=state.asset_id,
                            target_book=target_book,
                        )
                    )
                    continue

                raw, pl = client.process_transaction("transferAsset", params)

            # Check response
            status_code = str(pl.get("X_RETURN_STATUS") or "").strip()
            if status_code == "S":
                results.append(
                    TransferResult(
                        request_id=row.request_id,
                        asset_number=row.asset_number,
                        book_type_code=row.book_type_code,
                        mode="transfer",
                        status="TRANSFERRED",
                        asset_id=state.asset_id,
                        target_book=target_book,
                        fusion_response=pl,
                    )
                )
            else:
                results.append(
                    TransferResult(
                        request_id=row.request_id,
                        asset_number=row.asset_number,
                        book_type_code=row.book_type_code,
                        mode="transfer",
                        status="FAILED",
                        asset_id=state.asset_id,
                        target_book=target_book,
                        error=f"X_RETURN_STATUS={status_code}: {pl.get('X_MESSAGE_TEXT', '')}",
                        fusion_response=pl,
                    )
                )

        except Exception as exc:
            results.append(
                TransferResult(
                    request_id=row.request_id,
                    asset_number=row.asset_number,
                    book_type_code=row.book_type_code,
                    mode="transfer",
                    status="FAILED",
                    error=str(exc),
                )
            )

    return results


# ---------- Output ----------

def write_results(results: List[TransferResult], out_dir: Path) -> None:
    """Write results to JSON and CSV."""
    out_dir.mkdir(parents=True, exist_ok=True)

    # JSON
    results_json = [r.to_dict() for r in results]
    json_path = out_dir / "transfer_results.json"
    json_path.write_text(json.dumps(results_json, indent=2, default=str), encoding="utf-8")

    # CSV
    import csv
    csv_path = out_dir / "transfer_results.csv"
    if results_json:
        fieldnames = list(results_json[0].keys())
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in results_json:
                writer.writerow({k: json.dumps(v) if isinstance(v, dict) else v for k, v in row.items()})

    # Summary
    counts = {"total": len(results)}
    for r in results:
        counts[r.status] = counts.get(r.status, 0) + 1

    summary = {
        "run_ts": int(time.time()),
        "run_date": date.today().isoformat(),
        "counts": counts,
    }
    summary_path = out_dir / "transfer_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(f"\nResults written to: {out_dir}")
    print(f"  transfer_results.json  ({len(results)} rows)")
    print(f"  transfer_results.csv")
    print(f"  transfer_summary.json")


def print_summary(results: List[TransferResult]) -> None:
    counts: Dict[str, int] = {}
    for r in results:
        counts[r.status] = counts.get(r.status, 0) + 1

    print("\n" + "=" * 60)
    print("OFAM Excel Asset Transfer Summary")
    print("=" * 60)
    print(f"  Total rows processed : {len(results)}")
    for status, count in sorted(counts.items()):
        print(f"  {status:20s} : {count}")
    print("=" * 60)

    # Print failures
    failures = [r for r in results if r.status == "FAILED"]
    if failures:
        print(f"\nFailed transfers ({len(failures)}):")
        for f in failures:
            print(f"  - {f.request_id} ({f.asset_number}): {f.error}")


# ---------- CLI ----------

def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Read Excel asset transfer requests and execute via Oracle Fusion REST API."
    )
    parser.add_argument(
        "--excel",
        required=True,
        help="Path to the Excel file with asset transfer requests.",
    )
    parser.add_argument(
        "--config",
        default=str(REPO_ROOT / "sample_bip_with_pre_dff_updates.json"),
        help="Path to the JSON config file (oracle connection, field mappings, entity map).",
    )
    parser.add_argument(
        "--mode",
        choices=["dff", "transfer"],
        default="dff",
        help=(
            "'dff' = update DFF values only (processTransaction-updateAssetDescriptiveDetails). "
            "'transfer' = full transfer (getAssetInformation + transferAsset/bookTransfer)."
        ),
    )
    run_group = parser.add_mutually_exclusive_group(required=True)
    run_group.add_argument("--dry-run", action="store_true", help="Build payloads but don't call Oracle.")
    run_group.add_argument("--execute", action="store_true", help="Actually call Oracle Fusion APIs.")

    parser.add_argument(
        "--out-dir",
        default=str(REPO_ROOT / "output" / "excel_transfers"),
        help="Directory for output files.",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable debug logging.",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)-8s %(name)s - %(message)s",
    )

    # Read config
    config_path = Path(args.config)
    if not config_path.exists():
        print(f"ERROR: Config file not found: {config_path}", file=sys.stderr)
        return 1
    config = json.loads(config_path.read_text(encoding="utf-8"))

    # Read Excel
    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}", file=sys.stderr)
        return 1

    print(f"Reading Excel: {excel_path}")
    rows = read_excel(excel_path)
    print(f"Found {len(rows)} transfer request(s)")

    if not rows:
        print("No rows to process. Exiting.")
        return 0

    # Validate rows
    for row in rows:
        issues = []
        if not row.book_type_code:
            issues.append("missing BOOK_TYPE_CODE")
        if not row.transfer_to_entity:
            issues.append("missing TRANSFER_TO_ENTITY")
        if not row.transfer_date:
            issues.append("missing TRANSFER_DATE")
        if issues:
            print(f"  WARNING Row {row.row_number} ({row.request_id}): {', '.join(issues)}")

    dry_run = args.dry_run
    mode = args.mode
    print(f"Mode: {mode}, {'DRY-RUN' if dry_run else 'EXECUTE'}")

    # Run
    if mode == "dff":
        results = run_dff_mode(rows, config, dry_run)
    else:
        results = run_transfer_mode(rows, config, dry_run)

    # Output
    print_summary(results)
    write_results(results, Path(args.out_dir))

    # Print planned params for dry-run
    if dry_run:
        print("\nPlanned payloads:")
        for r in results:
            print(f"\n  [{r.request_id}] {r.asset_number} @ {r.book_type_code}")
            if r.planned_params:
                for k, v in r.planned_params.items():
                    print(f"    {k}: {v}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
